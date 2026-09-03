from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "nifty100.db"
PEER_GROUPS_PATH = (
    PROJECT_ROOT / "data" / "raw" / "peer_groups.xlsx"
)
OUTPUT_PATH = (
    PROJECT_ROOT / "output" / "peer_comparison.xlsx"
)


# ============================================================
# METRICS
# ============================================================

METRICS = [
    "ROE",
    "ROCE",
    "Net Profit Margin",
    "D/E",
    "FCF",
    "PAT CAGR 5yr",
    "Revenue CAGR 5yr",
    "EPS CAGR 5yr",
    "Interest Coverage",
    "Asset Turnover",
]


# ============================================================
# EXCEL COLORS
# ============================================================

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

BOLD_FONT = Font(
    bold=True
)


# ============================================================
# LOAD PEER GROUPS
# ============================================================

def load_peer_groups() -> pd.DataFrame:
    """Load peer-group assignment data."""

    if not PEER_GROUPS_PATH.exists():
        raise FileNotFoundError(
            f"Peer groups file not found: {PEER_GROUPS_PATH}"
        )

    df = pd.read_excel(
        PEER_GROUPS_PATH
    )

    required = {
        "peer_group_name",
        "company_id",
        "is_benchmark",
    }

    missing = required - set(df.columns)

    if missing:
        raise KeyError(
            f"Missing columns in peer_groups.xlsx: "
            f"{sorted(missing)}"
        )

    df = df[
        [
            "peer_group_name",
            "company_id",
            "is_benchmark",
        ]
    ].copy()

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
    )

    df["peer_group_name"] = (
        df["peer_group_name"]
        .astype(str)
        .str.strip()
    )

    df["is_benchmark"] = (
        df["is_benchmark"]
        .fillna(False)
        .astype(bool)
    )

    return df


# ============================================================
# LOAD PEER PERCENTILES
# ============================================================

def load_percentiles() -> pd.DataFrame:
    """Load Day 18 peer percentile table."""

    with sqlite3.connect(
        DB_PATH
    ) as conn:

        df = pd.read_sql_query(
            """
            SELECT
                company_id,
                peer_group_name,
                metric,
                value,
                percentile_rank,
                year
            FROM peer_percentiles
            ORDER BY
                peer_group_name,
                company_id
            """,
            conn,
        )

    if df.empty:
        raise ValueError(
            "peer_percentiles table is empty."
        )

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
    )

    return df


# ============================================================
# LOAD COMPANY NAMES
# ============================================================

def load_company_names() -> pd.DataFrame:
    """Load company names."""

    with sqlite3.connect(
        DB_PATH
    ) as conn:

        tables = pd.read_sql_query(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
            """,
            conn,
        )

        if "companies" not in tables[
            "name"
        ].tolist():
            return pd.DataFrame(
                columns=[
                    "company_id",
                    "company_name",
                ]
            )

        companies = pd.read_sql_query(
            "SELECT * FROM companies",
            conn,
        )

    if "id" in companies.columns:
        companies = companies.rename(
            columns={
                "id": "company_id"
            }
        )

    if "company_name" not in companies.columns:
        return pd.DataFrame(
            columns=[
                "company_id",
                "company_name",
            ]
        )

    companies["company_id"] = (
        companies["company_id"]
        .astype(str)
        .str.strip()
    )

    return (
        companies[
            [
                "company_id",
                "company_name",
            ]
        ]
        .drop_duplicates(
            "company_id"
        )
    )


# ============================================================
# BUILD WIDE PEER REPORT
# ============================================================

def build_group_dataframe(
    group_name: str,
    percentiles: pd.DataFrame,
    peers: pd.DataFrame,
    companies: pd.DataFrame,
) -> pd.DataFrame:
    """Build one wide dataframe for one peer group."""

    group = percentiles.loc[
        percentiles[
            "peer_group_name"
        ] == group_name
    ].copy()

    if group.empty:
        return pd.DataFrame()

    # --------------------------------------------------------
    # Metric values
    # --------------------------------------------------------

    values = group.pivot_table(
        index="company_id",
        columns="metric",
        values="value",
        aggfunc="first",
    )

    values.columns = [
        f"{metric}_value"
        for metric in values.columns
    ]

    # --------------------------------------------------------
    # Percentile ranks
    # --------------------------------------------------------

    ranks = group.pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank",
        aggfunc="first",
    )

    ranks.columns = [
        f"{metric}_percentile"
        for metric in ranks.columns
    ]

    # --------------------------------------------------------
    # Combine
    # --------------------------------------------------------

    result = values.join(
        ranks,
        how="outer",
    ).reset_index()

    # --------------------------------------------------------
    # Company names
    # --------------------------------------------------------

    result = result.merge(
        companies,
        on="company_id",
        how="left",
    )

    # --------------------------------------------------------
    # Benchmark flag
    # --------------------------------------------------------

    benchmark = peers.loc[
        peers[
            "peer_group_name"
        ] == group_name,
        [
            "company_id",
            "is_benchmark",
        ],
    ].drop_duplicates(
        "company_id"
    )

    result = result.merge(
        benchmark,
        on="company_id",
        how="left",
    )

    result["is_benchmark"] = (
        result["is_benchmark"]
        .fillna(False)
        .astype(bool)
    )

    # --------------------------------------------------------
    # Exact requested ordering
    # --------------------------------------------------------

    ordered_columns = [
        "company_id",
        "company_name",
    ]

    for metric in METRICS:

        value_column = (
            f"{metric}_value"
        )

        if value_column in result.columns:
            ordered_columns.append(
                value_column
            )

    for metric in METRICS:

        percentile_column = (
            f"{metric}_percentile"
        )

        if percentile_column in result.columns:
            ordered_columns.append(
                percentile_column
            )

    # Internal benchmark flag stays available.
    ordered_columns.append(
        "is_benchmark"
    )

    result = result[
        [
            column
            for column in ordered_columns
            if column in result.columns
        ]
    ]

    return result.sort_values(
        by="company_id"
    ).reset_index(
        drop=True
    )


# ============================================================
# ADD MEDIAN SUMMARY
# ============================================================

def add_summary_row(
    ws,
    dataframe: pd.DataFrame,
):
    """Add peer-group median row."""

    summary_row = ws.max_row + 2

    ws.cell(
        row=summary_row,
        column=1,
        value="Peer Group Median",
    )

    ws.cell(
        row=summary_row,
        column=1,
    ).fill = SUMMARY_FILL

    ws.cell(
        row=summary_row,
        column=1,
    ).font = BOLD_FONT

    header_map = {
        cell.value: cell.column
        for cell in ws[1]
    }

    for metric in METRICS:

        column_name = (
            f"{metric}_value"
        )

        if column_name not in header_map:
            continue

        values = pd.to_numeric(
            dataframe[column_name],
            errors="coerce",
        )

        if not values.notna().any():
            continue

        median_value = values.median()

        cell = ws.cell(
            row=summary_row,
            column=header_map[
                column_name
            ],
            value=float(median_value),
        )

        cell.fill = SUMMARY_FILL
        cell.font = BOLD_FONT


# ============================================================
# STYLE WORKSHEET
# ============================================================

def style_sheet(
    ws,
    dataframe: pd.DataFrame,
):
    """Apply Day 20 workbook formatting."""

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    for cell in ws[1]:

        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A2"

    # --------------------------------------------------------
    # Header map
    # --------------------------------------------------------

    header_map = {
        cell.value: cell.column
        for cell in ws[1]
    }

    # --------------------------------------------------------
    # Percentile colors
    # --------------------------------------------------------

    percentile_columns = [
        column_name
        for column_name in header_map
        if str(column_name).endswith(
            "_percentile"
        )
    ]

    for column_name in percentile_columns:

        column_index = header_map[
            column_name
        ]

        for row_index in range(
            2,
            ws.max_row + 1,
        ):

            cell = ws.cell(
                row=row_index,
                column=column_index,
            )

            if cell.value is None:
                continue

            try:
                percentile = float(
                    cell.value
                )
            except (
                TypeError,
                ValueError,
            ):
                continue

            if percentile >= 0.75:

                cell.fill = GREEN_FILL

            elif percentile >= 0.25:

                cell.fill = YELLOW_FILL

            else:

                cell.fill = RED_FILL

    # --------------------------------------------------------
    # Benchmark row
    #
    # dataframe still contains is_benchmark.
    # --------------------------------------------------------

    if "is_benchmark" in dataframe.columns:

        for dataframe_index, is_benchmark in enumerate(
            dataframe["is_benchmark"],
            start=2,
        ):

            if not bool(is_benchmark):
                continue

            # Only shade visible report columns.
            for column_index in range(
                1,
                ws.max_column + 1,
            ):

                cell = ws.cell(
                    row=dataframe_index,
                    column=column_index,
                )

                cell.fill = BENCHMARK_FILL

    # --------------------------------------------------------
    # Hide internal benchmark column
    # --------------------------------------------------------

    if "is_benchmark" in header_map:

        column_index = header_map[
            "is_benchmark"
        ]

        letter = ws.cell(
            row=1,
            column=column_index,
        ).column_letter

        ws.column_dimensions[
            letter
        ].hidden = True

    # --------------------------------------------------------
    # Column widths
    # --------------------------------------------------------

    for column_cells in ws.columns:

        max_length = 0

        for cell in column_cells:

            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value)),
            )

        width = min(
            max(
                max_length + 2,
                12,
            ),
            28,
        )

        ws.column_dimensions[
            column_cells[
                0
            ].column_letter
        ].width = width


# ============================================================
# GENERATE REPORT
# ============================================================

def generate_report():
    """Generate output/peer_comparison.xlsx."""

    print(
        "N100 Day 20 - Peer Comparison Report"
    )

    peers = load_peer_groups()

    percentiles = load_percentiles()

    companies = load_company_names()

    groups = sorted(
        percentiles[
            "peer_group_name"
        ]
        .dropna()
        .unique()
        .tolist()
    )

    print(
        f"Peer groups found: {len(groups)}"
    )

    if len(groups) != 11:
        raise ValueError(
            "Expected exactly 11 peer groups, "
            f"found {len(groups)}"
        )

    report_data = {}

    # --------------------------------------------------------
    # Build all 11 groups
    # --------------------------------------------------------

    for group_name in groups:

        dataframe = build_group_dataframe(
            group_name=group_name,
            percentiles=percentiles,
            peers=peers,
            companies=companies,
        )

        report_data[
            group_name
        ] = dataframe

        print(
            f"{group_name}: "
            f"{len(dataframe)} companies"
        )

    # --------------------------------------------------------
    # Write Excel
    # --------------------------------------------------------

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for group_name, dataframe in report_data.items():

            # Keep benchmark flag internally for
            # styling, then hide the column.
            dataframe.to_excel(
                writer,
                sheet_name=group_name[:31],
                index=False,
            )

    # --------------------------------------------------------
    # Style workbook
    # --------------------------------------------------------

    workbook = load_workbook(
        OUTPUT_PATH
    )

    for group_name, dataframe in report_data.items():

        worksheet = workbook[
            group_name[:31]
        ]

        style_sheet(
            worksheet,
            dataframe,
        )

        add_summary_row(
            worksheet,
            dataframe,
        )

    workbook.save(
        OUTPUT_PATH
    )

    print(
        "\nWorkbook created:"
    )

    print(
        OUTPUT_PATH
    )


# ============================================================
# MAIN
# ============================================================

def main():
    generate_report()


if __name__ == "__main__":
    main()