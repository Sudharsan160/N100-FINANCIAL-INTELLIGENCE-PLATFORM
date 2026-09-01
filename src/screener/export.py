from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.screener.engine import ScreenerEngine


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_PATH = PROJECT_ROOT / "output" / "screener_output.xlsx"


PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch",
]


# ------------------------------------------------------------
# Columns required by the Day 17 deliverable
# ------------------------------------------------------------

EXPORT_COLUMNS = [
    "company_id",
    "company_name",
    "year",
    "broad_sector",
    "sub_sector",
    "return_on_equity_pct",
    "roce_pct",
    "net_profit_margin_pct",
    "free_cash_flow_cr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "debt_to_equity",
    "interest_coverage",
    "operating_profit_margin_pct",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "market_cap_crore",
    "asset_turnover",
    "sales",
    "net_profit",
    "composite_quality_score",
]


# ------------------------------------------------------------
# Formatting
# ------------------------------------------------------------

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

HEADER_FONT = Font(
    bold=True,
)

WHITE_FONT = Font(
    color="000000",
)


def apply_header_format(ws):
    """Format header row."""

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_size_columns(ws):
    """Set readable column widths."""

    for column_cells in ws.columns:

        max_length = 0
        column_index = column_cells[0].column

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(
                max_length,
                len(str(value)),
            )

        width = min(
            max(max_length + 2, 12),
            28,
        )

        ws.column_dimensions[
            get_column_letter(column_index)
        ].width = width


# ------------------------------------------------------------
# Threshold colour coding
# ------------------------------------------------------------

def threshold_passes(
    value: Any,
    metric: str,
    thresholds: dict[str, Any],
) -> bool | None:
    """
    Return:
        True  -> meets threshold
        False -> fails threshold
        None  -> metric has no threshold in this preset
    """

    if metric == "roe_min":
        column = "return_on_equity_pct"
        return value >= thresholds["roe_min"]

    if metric == "de_max":
        column = "debt_to_equity"

        return value <= thresholds["de_max"]

    if metric == "fcf_min":
        return value > thresholds["fcf_min"]

    if metric == "revenue_cagr_5yr_min":
        return value >= thresholds["revenue_cagr_5yr_min"]

    if metric == "pat_cagr_5yr_min":
        return value >= thresholds["pat_cagr_5yr_min"]

    if metric == "opm_min":
        return value >= thresholds["opm_min"]

    if metric == "pe_max":
        return value <= thresholds["pe_max"]

    if metric == "pb_max":
        return value <= thresholds["pb_max"]

    if metric == "dividend_yield_min":
        return value >= thresholds["dividend_yield_min"]

    if metric == "icr_min":
        return value >= thresholds["icr_min"]

    if metric == "market_cap_min":
        return value >= thresholds["market_cap_min"]

    if metric == "net_profit_min":
        return value >= thresholds["net_profit_min"]

    if metric == "eps_cagr_min":
        return value >= thresholds["eps_cagr_min"]

    if metric == "asset_turnover_min":
        return value >= thresholds["asset_turnover_min"]

    if metric == "sales_min":
        return value >= thresholds["sales_min"]

    if metric == "dividend_payout_max":
        return value <= thresholds["dividend_payout_max"]

    if metric == "revenue_cagr_3yr_min":
        return value >= thresholds["revenue_cagr_3yr_min"]

    if metric == "de_exact":
        return abs(value - thresholds["de_exact"]) < 1e-9

    return None


FILTER_TO_COLUMN = {
    "roe_min": "return_on_equity_pct",
    "de_max": "debt_to_equity",
    "fcf_min": "free_cash_flow_cr",
    "revenue_cagr_5yr_min": "revenue_cagr_5yr",
    "pat_cagr_5yr_min": "pat_cagr_5yr",
    "opm_min": "operating_profit_margin_pct",
    "pe_max": "pe_ratio",
    "pb_max": "pb_ratio",
    "dividend_yield_min": "dividend_yield_pct",
    "icr_min": "interest_coverage",
    "market_cap_min": "market_cap_crore",
    "net_profit_min": "net_profit",
    "eps_cagr_min": "eps_cagr_5yr",
    "asset_turnover_min": "asset_turnover",
    "sales_min": "sales",
    "dividend_payout_max": "dividend_payout_ratio_pct",
    "revenue_cagr_3yr_min": "revenue_cagr_3yr",
    "de_exact": "debt_to_equity",
}


def colour_code_threshold_cells(
    ws,
    thresholds: dict[str, Any],
):
    """
    Apply:
        green = threshold met
        red   = threshold failed
    """

    header_map = {}

    for cell in ws[1]:
        header_map[cell.value] = cell.column

    for filter_name, threshold in thresholds.items():

        if filter_name not in FILTER_TO_COLUMN:
            continue

        column_name = FILTER_TO_COLUMN[
            filter_name
        ]

        if column_name not in header_map:
            continue

        column_index = header_map[
            column_name
        ]

        for row in range(2, ws.max_row + 1):

            cell = ws.cell(
                row=row,
                column=column_index,
            )

            value = cell.value

            if value is None:
                continue

            try:
                passed = threshold_passes(
                    float(value),
                    filter_name,
                    thresholds,
                )
            except (
                TypeError,
                ValueError,
            ):
                continue

            if passed is True:
                cell.fill = GREEN_FILL

            elif passed is False:
                cell.fill = RED_FILL


# ------------------------------------------------------------
# Export
# ------------------------------------------------------------

def build_screener_excel(
    engine: ScreenerEngine,
    output_path: Path = OUTPUT_PATH,
):
    """Run all presets and create screener_output.xlsx."""

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    preset_results: dict[
        str,
        pd.DataFrame,
    ] = {}

    print("Running screener presets...")

    for preset_name in PRESETS:

        result = engine.preset(
            preset_name
        )

        preset_results[
            preset_name
        ] = result

        display_name = (
            engine.config[
                "presets"
            ][preset_name].get(
                "display_name",
                preset_name,
            )
        )

        print(
            f"{display_name}: "
            f"{len(result)} companies"
        )

    # --------------------------------------------------------
    # Write Excel
    # --------------------------------------------------------

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        for preset_name, result in preset_results.items():

            display_name = (
                engine.config[
                    "presets"
                ][preset_name].get(
                    "display_name",
                    preset_name,
                )
            )

            export_columns = [
                column
                for column in EXPORT_COLUMNS
                if column in result.columns
            ]

            sheet_data = result[
                export_columns
            ].copy()

            sheet_data.to_excel(
                writer,
                sheet_name=display_name[:31],
                index=False,
            )

    # --------------------------------------------------------
    # Styling
    # --------------------------------------------------------

    workbook = load_workbook(
        output_path
    )

    for preset_name in PRESETS:

        display_name = (
            engine.config[
                "presets"
            ][preset_name].get(
                "display_name",
                preset_name,
            )
        )

        sheet_name = display_name[:31]

        ws = workbook[
            sheet_name
        ]

        apply_header_format(ws)
        auto_size_columns(ws)

        thresholds = (
            engine.config[
                "presets"
            ][preset_name].get(
                "thresholds",
                {},
            )
        )

        colour_code_threshold_cells(
            ws,
            thresholds,
        )

    workbook.save(
        output_path
    )

    return preset_results


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def main():

    print(
        "N100 Day 17 - Screener Excel Export"
    )

    engine = ScreenerEngine()

    results = build_screener_excel(
        engine
    )

    print(
        f"\nExcel export created:"
        f"\n{OUTPUT_PATH}"
    )

    print("\nPreset summary:")

    for preset_name, result in results.items():

        display_name = (
            engine.config[
                "presets"
            ][preset_name].get(
                "display_name",
                preset_name,
            )
        )

        print(
            f"{display_name}: "
            f"{len(result)} companies"
        )


if __name__ == "__main__":
    main()